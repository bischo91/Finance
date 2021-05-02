# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'helloworld_20200314.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!

import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import os
import glob
import csv
import openpyxl
from openpyxl import load_workbook
from xlsxwriter.workbook import Workbook
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import datetime


bank_dic = {'Chase': 2, 'Discover': 2}

def is_float_try(a):
    # check if 'a' is float and return boolean

    try:
        float(a)
        return True
    except TypeError:
        return False
    except ValueError:
        return False

def which_bank(file_name_input):
    # Determine which bank from file name and return bank name

    if 'CHASE' in file_name_input.upper() and '3225' in file_name_input:
        return 'Chase'
    elif 'DISCOVER' in file_name_input.upper() or 'DFS' in file_name_input.upper():
        return 'Discover'
    elif 'CHASE' in file_name_input.upper() and '7472':
        return 'Chase Credit'
    else:
        return 'Unknown'

def bank_condition(bank_name, desc, amt):
    # condition to which category/spending to count and returns boolean
    if bank_name == 'Chase':
        # for bank name chase, spending will not count if it's transaction with Robinhood, Discover, Mattrix, and Varo, or if amount < 0
        cond = ('DISCOVER' not in desc and 'ROBINHOOD' not in desc and 'MATTRIX' not in desc and 'Varo' not in desc and 'CHASE CREDIT' not in desc.upper())
        cond = cond and float(amt)<0
    elif bank_name == 'Discover':
        # for bank name discover, spending will not count if it's Internet Payment or Cashback bonus
        cond = ('INTERNET PAYMENT' not in desc) and ('CASHBACK BONUS' not in desc)
    return cond

def find_category_chase(bank_name, desc, amt):
    # return category of spending for Chase
    if bank_name == 'Chase':
        if 'RENT' in desc.upper() or 'GRU' in desc.upper() or 'TIOGA' in desc.upper() or 'CLAY' in desc.upper():
            cat_spend_chase = 'Rent / Utility'
        elif 'AIR' in desc.upper():
            cat_spend_chase = 'Travel/ Entertainment'
        elif 'MOBILE' in desc.upper():
            cat_spend_chase = 'Phone/ Wifi'
        elif 'COSTCO' in desc.upper() or 'WAL-MART' in desc.upper():
            cat_spend_chase = 'Grocery'
        elif 'COSTCO GAS' in desc.upper():
            cat_spend_chase = 'Gasoline'
        elif 'TARGET' in desc.upper():
            cat_spend_chase = 'Target'
        elif 'YOGA' in desc.upper():
            cat_spend_chase = 'Gym'
        else:
            cat_spend_chase = 'Unknown'
    return cat_spend_chase

def find_category_chase_credit(bank_name, desc, cat, amt):
    # return category of spending for Chase
    if bank_name == 'Chase Credit':
        if 'MOBILE' in desc.upper():
            cat_spend_chase = 'Phone/ Wifi'
        elif 'COSTCO' in desc.upper() or 'WAL-MART' in desc.upper() or cat == 'Groceries':
            cat_spend_chase = 'Grocery'
        elif 'COSTCO GAS' in desc.upper() or 'GAS' in cat.upper():
            cat_spend_chase = 'Gasoline'
            if float(amt) < 10:
                cat_spend_chase = 'BC'
        elif 'TARGET' in desc.upper():
            cat_spend_chase = 'Target'
        elif 'YOGA' in desc.upper() or 'GOLF' in desc.upper():
            cat_spend_chase = 'Gym'
        elif 'FOOD' in cat.upper():
            cat_spend_chase = 'Restaurant'
        elif 'SHOPPING' in cat.upper():
            cat_spend_chase = 'Merchandise'
        elif 'BILLS' in cat.upper():
            cat_spend_chase = 'Utilities'
        else:
            cat_spend_chase = cat
    return cat_spend_chase

def find_category_discover(bank_name, desc, cat, amt):
    # return category of spending for Discover
    if bank_name == 'Discover':
        if cat == 'Supermarkets':
            cat_return = 'Grocery'
        elif cat == 'Gasoline' and float(amt) < 10:
            cat_return = 'BC'
        elif 'ATHLETIC' in desc.upper():
            cat_return = 'Gym'
        elif cat == 'Warehouse Clubs':
            cat_return = 'Grocery'
        elif 'ATT' in desc:
            cat_return = 'Rent / Utility'
        elif 'WALMART' in desc.upper():
            cat_return = 'Grocery'
        elif cat == 'Payments and Credits':
            cat_return = 'Refund'
        elif 'MOBILE' in desc.upper():
            cat_return = 'Phone/ Wifi'
        else:
            cat_return = cat
    return cat_return

def month_to_num(month_str):
    # convert string to number for month
    switcher = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }
    return switcher.get(month_str)

def num_to_month(num):
    # convert number to string for month
    switcher = {
        1: "Jan",
        2: "Feb",
        3: "Mar",
        4: "Apr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Aug",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dec",
    }
    return switcher.get(num)

def num_to_month_full(num):
    # convert number to string for month
    switcher = {
        1: "January",
        2: "Febbruary",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December",
    }
    return switcher.get(num)

def spending(path_abs, month_selected, year_selected, only_saved_data):
    # read excel file for selected month and year
    # categorize each spending and calculate total spending
    # combine all valid spendings to return
    # input: load excel files in a folder (path_abs), month_selected, year_selected
    # return amount, category, description, bank, and date as each list, and total spending

    file_list = os.listdir(path_abs)
    total_spend_saved = 0
    amt_comb_saved = []
    cat_comb_saved = []
    desc_comb_saved = []
    bank_comb_saved = []
    date_comb_saved = []
    total_spend = 0
    amt_comb = []
    cat_comb = []
    desc_comb = []
    bank_comb = []
    date_comb = []

    for file_name_raw in file_list:
        # read files from folder
        if os.path.splitext(file_name_raw)[1] == '.xlsx':
            bank_name = which_bank(file_name_raw)
            file_name = path_abs + '/' + file_name_raw
            wb = load_workbook(filename=file_name, data_only=True)
            ws = wb.active

            if ws['A1'].value == 'Saved from program':
                # if data is from saved Excel
                j = 2
                while is_float_try(ws['C' + str(j)].value):
                    date_spent = ws['A' + str(j)].value
                    if int(date_spent[:2]) == month_selected and int(date_spent[-4:]) == year_selected:
                        date_comb_saved += [date_spent]
                        total_spend_saved += float(ws['C' + str(j)].value)
                        amt_comb_saved += [ws['C' + str(j)].value]
                        desc_comb_saved += [ws['D' + str(j)].value]
                        bank_comb_saved += [ws['E' + str(j)].value]
                        cat_comb_saved += [ws['B' + str(j)].value]
                    j += 1
            elif bank_name == 'Chase' or bank_name == 'Discover':
                # if data is from raw Excel (Chase/Discover)
                # bank define
                i = 2
                while is_float_try(ws['D' + str(i)].value):
                    if bank_name == 'Discover':
                        date_spent = ws['A' + str(i)].value
                    # assuming Chase if not Discover
                    elif bank_name == 'Chase':
                        date_spent = ws['B' + str(i)].value

                    if int(date_spent[:2]) == month_selected and int(date_spent[-4:]) == year_selected:
                        if 'Payment to Chase' not in ws['C' + str(i)].value:
                            # condition for date range
                            desc = ws['C' + str(i)].value
                            amt = ws['D' + str(i)].value
                            if bank_condition(bank_name, desc, amt):
                                # for Discover
                                if bank_name == 'Discover':
                                    cat = [find_category_discover(bank_name, desc, ws['E' + str(i)].value,  ws['D' + str(i)].value)]
                                # for Chase
                                else:
                                    amt = -float(ws['D' + str(i)].value)
                                    cat = [find_category_chase(bank_name, desc, amt)]
                                date_comb += [date_spent]
                                total_spend += float(amt)
                                amt_comb += [amt]
                                desc_comb += [ws['C' + str(i)].value]
                                if 'Target' in cat:
                                    cat_comb += ['Grocery']
                                    bank_comb += ['Target']
                                else:
                                    cat_comb += cat
                                    bank_comb += [bank_name]
                    i += 1

            elif bank_name == 'Chase Credit':
                k = 2
                while is_float_try(ws['F' + str(k)].value):
                    date_spent = ws['A' + str(k)].value
                    if int(date_spent[:2]) == month_selected and int(date_spent[-4:]) == year_selected:
                        # condition for date range
                        if -float(ws['F' + str(k)].value)>0:
                            desc = ws['C' + str(k)].value
                            cat = ws['D' + str(k)].value
                            amt = -float(ws['F' + str(k)].value)
                            cat = [find_category_chase_credit(bank_name, desc, cat, amt)]
                            date_comb += [date_spent]
                            total_spend += float(amt)
                            amt_comb += [amt]
                            desc_comb += [ws['C' + str(k)].value]
                            if 'Target' in cat:
                                cat_comb += ['Grocery']
                                bank_comb += ['Target']
                            else:
                                cat_comb += cat
                                bank_comb += [bank_name]
                    k += 1
                # print(desc_comb)

    if only_saved_data == True:
        # to read only saved data (for yearly spending)
        return  total_spend_saved, amt_comb_saved, cat_comb_saved, desc_comb_saved, bank_comb_saved, date_comb_saved
    if total_spend_saved == 0 and total_spend > 0:
        # if no saved data, read raw data automatically
        return total_spend, amt_comb, cat_comb, desc_comb, bank_comb, date_comb
    elif total_spend == 0 and total_spend_saved >0:
        # if no raw data, read saved data automatically
        return total_spend_saved, amt_comb_saved, cat_comb_saved, desc_comb_saved, bank_comb_saved, date_comb_saved
    elif total_spend > 0 and total_spend_saved >0:
        # if both raw and saved data exist, ask
        if QMessageBox.Yes == QMessageBox.question(None, 'Existing Saved Data', 'Do you want to load saved data for ' + num_to_month(month_selected) + '? [click no for raw data]', QMessageBox.Yes | QMessageBox.No):
            return total_spend_saved, amt_comb_saved, cat_comb_saved, desc_comb_saved, bank_comb_saved, date_comb_saved
        else:
            return total_spend, amt_comb, cat_comb, desc_comb, bank_comb, date_comb
    else:
        return 0, [], [], [], [], []

def unique(list_x):
    # input list, and return a list with unique elements
    unique_list = []
    for x in list_x:
        if x not in unique_list:
            unique_list.append(x)
    return unique_list

class Ui_Dialog(QWidget):
    def push_button_clicked(self):
        # locate folder and convert csv to xlsx and save
        file_name = QFileDialog.getExistingDirectory(self, "select Directory")
        if file_name != '':
            self.lineEdit.setText(file_name)
        for csvfile in glob.glob(os.path.join(self.lineEdit.text(), '*.csv')):
            workbook = Workbook(csvfile[:-4] + '.xlsx')
            worksheet = workbook.add_worksheet()
            with open(csvfile, 'rt', encoding='utf8') as f:
                reader = csv.reader(f)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        worksheet.write(r, c, col)
            workbook.close()
        self.update_table(self.lineEdit.text())

    def update_table(self, path_abs):
        # update table when date changed or file loaded
        month = int(month_to_num(self.comboBox.currentText()))
        year = int(self.dateEdit.date().year())
        self.tableWidget.horizontalHeader().setVisible(False)
        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget_2.horizontalHeader().setVisible(False)
        self.tableWidget_2.verticalHeader().setVisible(False)
        if path_abs != '': # path not empty
            self.total_spend, self.amt_comb, self.cat_comb, self.desc_comb, self.bank_comb, self.date_comb = spending(path_abs, month, year, False)
            if self.total_spend == 0:
                # no data loaded
                self.tableWidget.setRowCount(1)
                self.tableWidget.setColumnCount(1)
                self.tableWidget.setItem(0, 0, QTableWidgetItem('No Data found for selected month and year'))
                self.tableWidget_2.setRowCount(1)
                self.tableWidget_2.setColumnCount(1)
                self.tableWidget_2.setItem(0, 0, QTableWidgetItem(''))
                self.label_4.setText("")
                self.fig.set_visible(False)
                self.canvas.draw()
            else:
                # self.monthly_spending()
                self.length = len(self.amt_comb)
                self.tableWidget.setRowCount(self.length)
                self.tableWidget.setColumnCount(4)
                self.tableWidget.horizontalHeader().setVisible(True)
                self.tableWidget.setHorizontalHeaderLabels(['Amount ($)', 'Category', 'Description', 'Bank'])
                self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
                self.cat_combobox_all = []
                self.cat_list = unique(self.cat_comb)
                if 'BC' not in self.cat_comb:
                    self.cat_list.append('BC')
                elif 'JL' not in self.cat_comb:
                    self.cat_list.append('JL')
                for i in range(self.length):
                    # update table by each row
                    self.cat_combobox = QComboBox()
                    self.cat_combobox_all.append(self.cat_combobox)
                    self.cat_combobox.addItems(self.cat_list)
                    cat = self.cat_comb[i]
                    amt = float(self.amt_comb[i])
                    desc = self.desc_comb[i]
                    bank = self.bank_comb[i]
                    self.cat_combobox.setCurrentIndex(self.cat_list.index(cat))
                    self.tableWidget.setCellWidget(i, 1, self.cat_combobox)
                    self.tableWidget.setItem(i, 0, QTableWidgetItem(str("{:.2f}".format(amt))))
                    self.tableWidget.setItem(i, 2, QTableWidgetItem(desc))
                    self.tableWidget.setItem(i, 3, QTableWidgetItem(bank))
                    self.cat_combobox_all[i].currentIndexChanged.connect(self.cat_update)
                self.pushButton_2.setVisible(True)
                self.pushButton_3.setVisible(True)
                self.pushButton_4.setVisible(True)
                self.switch = True
                self.edit_switch()
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget_2.resizeColumnsToContents()
        self.tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget.setMouseTracking(False)

    def cat_selected_cal(self):
        total_selected = 0
        self.cat_selected = []
        self.cost_cat = [0] * len(self.cat_pool)
        bc_spent = 0
        jl_spent = 0
        if self.radiobutton_bank_1.isChecked():
            bank_selected = 'total'
        elif self.radiobutton_bank_2.isChecked():
            bank_selected = 'chase'
        elif self.radiobutton_bank_3.isChecked():
            bank_selected = 'discover'
        elif self.radiobutton_bank_4.isChecked():
            bank_selected = 'target'
        elif self.radiobutton_bank_5.isChecked():
            bank_selected = 'chase credit'
        for i in range(len(self.cat_pool)):
            if self.checkbox_all[i].isChecked():
                self.cat_selected.append(self.cat_pool[i])
        for j in range(self.length):
            if self.bank_comb[j].lower() == bank_selected or bank_selected == 'total':
                if self.cat_combobox_all[j].currentText() in self.cat_selected:
                    total_selected += float(self.amt_comb[j])
                k = self.cat_pool.index(self.cat_combobox_all[j].currentText())
                self.cost_cat[k] += float(self.amt_comb[j])
        for l in range(len(self.cat_pool)):
            if self.cat_pool[l] == "BC":
                bc_spent = self.cost_cat[l]
            elif self.cat_pool[l] == "JL":
                jl_spent = self.cost_cat[l]
            if self.cat_pool[l] in self.cat_selected and total_selected != 0:
                percent = 100*self.cost_cat[l]/total_selected
            else:
                percent = 0
            self.tableWidget_2.setItem(l, 1, QTableWidgetItem(self.cat_pool[l]))
            self.tableWidget_2.setItem(l, 2, QTableWidgetItem(str("{:.2f}".format(self.cost_cat[l]))))
            self.tableWidget_2.setItem(l, 3, QTableWidgetItem(str("{:.2f}".format(percent))))
        budget = 2361.47
        saving = budget - (self.total_spend - bc_spent - jl_spent)
        # bc = 333.03 - BC_spent
        bc = saving/2 + 333.03
        jl_total = bc
        bc_total = bc + 1177.37
        extra_print = "\n                         "
        if saving < 0:
            extra_print = extra_print + "- $" + str("{:.2f}".format(-saving))
        else:
            extra_print = extra_print + "+ $" + str("{:.2f}".format(saving))
        # "Monthly Income: $4204.90 \n
        self.label_4.setText("Selected Spending: $" + str("{:.2f}".format(total_selected)) \
                             +"\n\nBudget:            + $2361.47" \
                             + "\n  Spending:         - $" + str("{:.2f}".format(self.total_spend - bc_spent - jl_spent)) \
                             + "\n-----------------------------------"
                             + extra_print \
                             + "\n\n\nBrian's spending: $" + str("{:.2f}".format(bc_spent)) \
                             + "\nBrian receives $" + str("{:.2f}".format(bc_total - bc_spent))
                             +"\n\nJessica's spending: $" + str("{:.2f}".format(jl_spent))
                             +"\nJessica receives $" + str("{:.2f}".format(jl_total - jl_spent)))
        self.tableWidget_2.resizeColumnsToContents()
        if total_selected > 0:
            self.plot_graph()
            self.fig.set_visible(True)
            self.canvas.draw()
        else:
            self.fig.set_visible(False)
            self.canvas.draw()

    def plot_graph(self):
        ratio = []
        labels = []
        for i in range(len(self.cat_pool)):
            if self.cat_pool[i] in self.cat_selected and float(self.tableWidget_2.item(i,3).text()) > 0:
                ratio.append(self.tableWidget_2.item(i,3).text())
                labels.append(self.cat_pool[i])
        self.ax.cla()
        class pieclass:
            i = -1
        def pielabel(pct, labels):
            pieclass.i +=1
            return "{:s}\n{:.2f} %".format(labels[pieclass.i], pct)
        self.ax.pie(ratio, autopct=lambda pct: pielabel(pct, labels), pctdistance=0.75, radius=1.55)
        self.ax.set_visible(True)
        self.canvas.draw()

    def cat_cal(self):
        self.cat_pool = []
        self.cost_cat = []
        for k in range(self.length):
            cat = self.cat_combobox_all[k].currentText()
            amt = float(self.amt_comb[k])
            if cat in self.cat_pool:
                index = self.cat_pool.index(cat)
                self.cost_cat[index] += amt
            else:
                self.cat_pool += [cat]
                self.cost_cat.append(amt)

    def cat_update(self):
        self.cat_cal()
        length_cat = len(self.cat_pool)
        self.tableWidget_2.setRowCount(length_cat)
        self.tableWidget_2.setColumnCount(4)
        self.tableWidget_2.horizontalHeader().setVisible(True)
        self.tableWidget_2.setHorizontalHeaderLabels(['', 'Category', 'Amount ($)', '%'])
        self.radiobutton_bank_1.setVisible(True)
        self.radiobutton_bank_1.setChecked(True)
        self.radiobutton_bank_2.setVisible(True)
        self.radiobutton_bank_3.setVisible(True)
        self.radiobutton_bank_4.setVisible(True)
        self.radiobutton_bank_5.setVisible(True)
        self.checkbox_all = []
        for j in range(length_cat):
            # percent = 100*self.cost_cat[j] / sum(self.cost_cat)
            self.checkbox = QCheckBox()
            self.checkbox.setChecked(True)
            self.checkbox_all.append(self.checkbox)
            self.tableWidget_2.setCellWidget(j, 0, self.checkbox)
            self.checkbox_all[j].stateChanged.connect(self.cat_selected_cal)
        self.cat_selected_cal()

    def new_cat_button_clicked(self):
        text, ok = QInputDialog().getText(self, "Adding New Category", "Enter Category:")
        combobox_list = []
        if ok and text:
            error = 0
            while error == 0:
                for i in range(len(self.cat_combobox)):
                    combobox_list.append(self.cat_combobox.itemText(i))
                if text.lower() not in [x.lower() for x in combobox_list]:
                    for j in range(self.length):
                        self.cat_combobox_all[j].addItem(str(text)[0].upper() + str(text)[1::])
                    error += 1
                else:
                    text, ok = QInputDialog().getText(self, "Adding New Category", "Error: it's alreaedy there tho")

    def date_changed(self):
        self.update_table(self.lineEdit.text())

    def edit_switch(self):
        if self.switch is False:
            self.switch = True
            self.pushButton_2.setText("Edit Enabled")
        else:
            self.switch = False
            self.pushButton_2.setText("Edit Disabled")
        for i_switch in range(len(self.cat_combobox_all)):
            self.cat_combobox_all[i_switch].setEnabled(self.switch)
        self.cat_update()

    def monthly_spending(self):
        path_abs = self.lineEdit.text()
        if path_abs != '':
            month_from = int(month_to_num(self.comboBox_2.currentText()))
            month_to = int(month_to_num(self.comboBox_3.currentText()))
            year_from = int(self.dateEdit_3.date().year())
            year_to = int(self.dateEdit_4.date().year())
            monthly_spend = []
            month_str = []
            if year_to < year_from:
                self.label_8.setText("Error: Year Error")
                self.fig_2.set_visible(False)
            elif year_to > year_from:
                self.fig_2.set_visible(True)
                self.label_8.setText("")
                # first year
                for h in range(month_from, 13):
                    monthly_spend.append(spending(path_abs, h, year_from, True)[0])
                    month_str.append(num_to_month(h) + "\n %d" % (year_from-2000))
                # mid year
                if year_to != year_from:
                    for i in range(year_from+1, year_to):
                        for j in range(12):
                            monthly_spend.append(spending(path_abs, j+1, i, True)[0])
                            month_str.append(num_to_month(j+1) + "\n %d" % (i-2000))
                    # last year
                    for k in range(1, month_to+1):
                        monthly_spend.append(spending(path_abs, k, year_to, True)[0])
                        month_str.append(num_to_month(k) + "\n %d" % (year_to-2000))
            else:
                if month_from > month_to:
                    self.label_8.setText("Error: Month Error")
                    self.fig_2.set_visible(False)
                else:
                    self.label_8.setText("")
                    self.fig_2.set_visible(True)
                    for l in range(month_from, month_to+1):
                        monthly_spend.append(spending(path_abs, l, year_to, True)[0])
                        month_str.append(num_to_month(l) + "\n %d" % (year_to-2000))
            self.ax_2.cla()
            self.ax_2.bar(np.arange(len(monthly_spend)), monthly_spend)
            plt.xticks(np.arange(len(monthly_spend)), tuple(month_str))
            plt.ylabel("Total Spending ($)")
            self.canvas_2.draw()

    def save_data(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value = 'Saved from program'
        for i in range(self.length):
            j = i+2
            sheet.cell(row=j, column=1).value = self.date_comb[i]
            sheet.cell(row=j, column=2).value = self.cat_combobox_all[i].currentText()
            sheet.cell(row=j, column=3).value = self.amt_comb[i]
            sheet.cell(row=j, column=4).value = self.desc_comb[i]
            sheet.cell(row=j, column=5).value = self.bank_comb[i]
        month = self.comboBox.currentText()
        year = int(self.dateEdit.date().year())
        wb.save(self.lineEdit.text() + '/' + month +'_' + str(year) +'.xlsx')

    def setupUi(self, Dialog):
        Dialog.setObjectName("Account Book")
        Dialog.setEnabled(True)
        Dialog.resize(1200, 750)
        # Vertical Layout
        self.verticalLayout = QtWidgets.QVBoxLayout(Dialog)
        self.verticalLayout.setObjectName("verticalLayout")
        # Frame 2
        self.frame_2 = QtWidgets.QFrame(Dialog)
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.frame_2)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalLayout.addWidget(self.frame_2)
        # Label in Frame 2
        self.label = QtWidgets.QLabel(self.frame_2)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        # LineEdit in Frame 2
        self.lineEdit = QtWidgets.QLineEdit(self.frame_2)
        self.lineEdit.setReadOnly(True)
        self.lineEdit.setObjectName("lineEdit")
        self.horizontalLayout.addWidget(self.lineEdit)
        # PushButton in Frame_2
        self.pushButton = QtWidgets.QPushButton(self.frame_2)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        # PushButton ------------------------------------------------------------------Action: Load
        self.pushButton.clicked.connect(self.push_button_clicked)
        # TableWidget
        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        # Frame 3
        self.frame_3 = QtWidgets.QFrame(self.tab)
        self.frame_3.setGeometry(QtCore.QRect(170, 10, 1000, 650))
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        # TableWidget in Frame 3
        self.tableWidget = QtWidgets.QTableWidget(self.frame_3)
        self.tableWidget.setGeometry(QtCore.QRect(10, 10, 625, 600))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        # TableWidget in Frame 3
        self.tableWidget_2 = QtWidgets.QTableWidget(self.frame_3)
        self.tableWidget_2.setGeometry(QtCore.QRect(645, 10, 280, 250))
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(0)
        self.tableWidget_2.setRowCount(0)
        # verticalLayoutWidget in Frame_3
        self.fig = plt.figure()
        plt.rc('font', size = 8)
        self.canvas = FigureCanvas(self.fig)
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.frame_3)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(635, 280, 360, 360))
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.addWidget(self.canvas)
        self.verticalLayout_2.setContentsMargins(0,0,0,0)
        self.ax = self.fig.add_subplot(111)
        self.fig.set_visible(False)
        # pushButton_2 in Frame_3
        self.pushButton_2 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_2.setGeometry(QtCore.QRect(10, 615, 120, 25))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.setVisible(False)
        self.pushButton_2.clicked.connect(self.edit_switch)
        # PushButton_3 in Frame_3
        self.pushButton_3 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_3.setGeometry(QtCore.QRect(150, 615, 120, 25))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.setVisible(False)
        self.pushButton_3.clicked.connect(self.new_cat_button_clicked)
        # PushButton_4 in Frame_3
        self.pushButton_4 = QtWidgets.QPushButton(self.frame_3)
        self.pushButton_4.setGeometry(QtCore.QRect(290, 615, 120, 25))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.setVisible(False)
        self.pushButton_4.clicked.connect(self.save_data)
        # Frame_8 in Frame_3
        self.frame_8 = QtWidgets.QFrame(self.frame_3)
        self.frame_8.setGeometry(QtCore.QRect(930, 10, 60, 270))
        self.frame_8.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_8.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_8.setObjectName("frame_8")
        # radiobutton in Frame_8
        self.radiobutton_bank_1 = QtWidgets.QRadioButton(self.frame_8)
        self.radiobutton_bank_1.setGeometry(QtCore.QRect(0, 10, 70, 20))
        self.radiobutton_bank_1.setText("Total")
        self.radiobutton_bank_1.setVisible(False)
        self.radiobutton_bank_1.clicked.connect(self.cat_selected_cal)
        self.radiobutton_bank_2 = QtWidgets.QRadioButton(self.frame_8)
        self.radiobutton_bank_2.setGeometry(QtCore.QRect(0, 50, 70, 20))
        self.radiobutton_bank_2.setText("Chase")
        self.radiobutton_bank_2.setVisible(False)
        self.radiobutton_bank_2.clicked.connect(self.cat_selected_cal)
        self.radiobutton_bank_3 = QtWidgets.QRadioButton(self.frame_8)
        self.radiobutton_bank_3.setGeometry(QtCore.QRect(0, 90, 70, 20))
        self.radiobutton_bank_3.setText("Discover")
        self.radiobutton_bank_3.setVisible(False)
        self.radiobutton_bank_3.clicked.connect(self.cat_selected_cal)
        self.radiobutton_bank_4 = QtWidgets.QRadioButton(self.frame_8)
        self.radiobutton_bank_4.setGeometry(QtCore.QRect(0, 130, 70, 20))
        self.radiobutton_bank_4.setText("Target")
        self.radiobutton_bank_4.setVisible(False)
        self.radiobutton_bank_4.clicked.connect(self.cat_selected_cal)
        self.radiobutton_bank_5 = QtWidgets.QRadioButton(self.frame_8)
        self.radiobutton_bank_5.setGeometry(QtCore.QRect(0, 170, 70, 20))
        self.radiobutton_bank_5.setText("Chase Credit")
        self.radiobutton_bank_5.setVisible(False)
        self.radiobutton_bank_5.clicked.connect(self.cat_selected_cal)
        # Frame 4
        self.frame_4 = QtWidgets.QFrame(self.tab)
        self.frame_4.setGeometry(QtCore.QRect(10, 10, 150, 650))
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        # Frame in Frame 4
        self.frame = QtWidgets.QFrame(self.frame_4)
        # self.frame.setGeometry(QtCore.QRect(10, 10, 130, 130))
        self.frame.setGeometry(QtCore.QRect(10, 10, 150, 130))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        # Label 4 in Frame 4
        self.label_4 = QtWidgets.QLabel(self.frame_4)
        self.label_4.setGeometry(QtCore.QRect(10, 175, 140, 300))
        self.label_4.setObjectName("label_4")
        self.label_4.setText("")
        # GridLayout_2 in Frame
        self.gridLayout_2 = QtWidgets.QGridLayout(self.frame)
        self.gridLayout_2.setObjectName("gridLayout_2")
        # Label_3 in GridLayout_2 in Frame
        self.label_3 = QtWidgets.QLabel(self.frame)
        self.label_3.setObjectName("label_3")
        self.gridLayout_2.addWidget(self.label_3, 1, 0, 1, 1)
        # DateEdit in GridLayout_2 in Frame
        self.dateEdit = QtWidgets.QDateEdit(self.frame)
        self.dateEdit.setDate(QtCore.QDate(datetime.date.today().year, 1, 1))
        self.dateEdit.setObjectName("dateEdit")
        self.gridLayout_2.addWidget(self.dateEdit, 0, 1, 1, 1)
        # DateEdit -------------------------------------------------------------------Action: Year Change
        self.dateEdit.dateChanged.connect(self.date_changed)
        # Label_2 in GridLayout_2 in Frame
        self.label_2 = QtWidgets.QLabel(self.frame)
        self.label_2.setObjectName("label_2")
        self.gridLayout_2.addWidget(self.label_2, 0, 0, 1, 1)
        # ComboBox in GridLayout_2 in Frame
        self.comboBox = QtWidgets.QComboBox(self.frame)
        self.comboBox.setEnabled(True)
        self.comboBox.setMouseTracking(False)
        self.comboBox.setFocusPolicy(QtCore.Qt.NoFocus)
        self.comboBox.setAcceptDrops(True)
        self.comboBox.setAutoFillBackground(False)
        self.comboBox.setEditable(False)
        self.comboBox.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("January")
        self.comboBox.addItem("February")
        self.comboBox.addItem("March")
        self.comboBox.addItem("April")
        self.comboBox.addItem("May")
        self.comboBox.addItem("June")
        self.comboBox.addItem("July")
        self.comboBox.addItem("August")
        self.comboBox.addItem("September")
        self.comboBox.addItem("October")
        self.comboBox.addItem("November")
        self.comboBox.addItem("December")
        self.comboBox.setCurrentText(num_to_month_full(datetime.date.today().month))
        self.comboBox.adjustSize()
        self.gridLayout_2.addWidget(self.comboBox, 1, 1, 1, 1)
        # ComboBox -------------------------------------------------------------------Action: Month Change
        self.comboBox.currentTextChanged.connect(self.date_changed)
        # Tab 2
        self.tabWidget.addTab(self.tab, "")
        self.tab2 = QtWidgets.QWidget()
        self.tab2.setObjectName("tab2")
        self.tabWidget.addTab(self.tab2, "")
        self.verticalLayout.addWidget(self.tabWidget)
        # Frame 6 in Tab 2
        self.frame_6 = QtWidgets.QFrame(self.tab2)
        self.frame_6.setGeometry(QtCore.QRect(10, 10, 1155, 645))
        self.frame_6.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_6.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_6.setObjectName("frame_6")
        # Frame 5 in Tab 2
        self.frame_5 = QtWidgets.QFrame(self.frame_6)
        self.frame_5.setGeometry(QtCore.QRect(20, 10, 220, 40))
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        # HorizontalLayout_2 in Frame 5
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.frame_5)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        # Label_5 in Frame 5
        self.label_5 = QtWidgets.QLabel(self.frame_5)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_2.addWidget(self.label_5)
        # Label_6 in Frame_5
        self.label_6 = QtWidgets.QLabel(self.frame_5)
        self.label_6.setObjectName("label_6")
        self.horizontalLayout_2.addWidget(self.label_6)
        # DateEdit_3 in Frame_5
        self.dateEdit_3 = QtWidgets.QDateEdit(self.frame_5)
        self.dateEdit_3.setDate(QtCore.QDate(datetime.date.today().year, 1, 1))
        self.dateEdit_3.setObjectName("dateEdit_3")
        self.horizontalLayout_2.addWidget(self.dateEdit_3)
        # ComboBox_2 in Frame_5
        self.comboBox_2 = QtWidgets.QComboBox(self.frame_5)
        self.comboBox_2.setObjectName("comboBox_2")
        self.horizontalLayout_2.addWidget(self.comboBox_2)
        self.comboBox_2.addItem("January")
        self.comboBox_2.addItem("February")
        self.comboBox_2.addItem("March")
        self.comboBox_2.addItem("April")
        self.comboBox_2.addItem("May")
        self.comboBox_2.addItem("June")
        self.comboBox_2.addItem("July")
        self.comboBox_2.addItem("August")
        self.comboBox_2.addItem("September")
        self.comboBox_2.addItem("October")
        self.comboBox_2.addItem("November")
        self.comboBox_2.addItem("December")
        self.comboBox_2.setCurrentText(num_to_month_full(datetime.date.today().month))
        self.comboBox_2.adjustSize()
        # fig_2 in Frame_6
        self.fig_2 = plt.figure()
        self.ax_2 = self.fig_2.add_subplot(111)
        self.canvas_2 = FigureCanvas(self.fig_2)
        self.verticalLayoutWidget_3 = QtWidgets.QWidget(self.frame_6)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(-75, 40, 1300, 600))
        # Error message in Frame_6
        self.label_8 = QtWidgets.QLabel(self.frame_6)
        self.label_8.setObjectName("label_8")
        self.label_8.setGeometry(QtCore.QRect(100, 60, 300, 20))

        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.verticalLayout_3.addWidget(self.canvas_2)
        self.verticalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.fig_2.set_visible(False)
        # Frame_7 in Frame_6
        self.frame_7 = QtWidgets.QFrame(self.frame_6)
        self.frame_7.setGeometry(QtCore.QRect(350, 10, 220, 40))
        self.frame_7.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_7.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_7.setObjectName("frame_7")
        # HorizontalLayout_3 in Frame_7
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout(self.frame_7)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        # Label_7 in Frame_7
        self.label_7 = QtWidgets.QLabel(self.frame_7)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout_3.addWidget(self.label_7)
        # DateEdit_4 in Frame_7
        self.dateEdit_4 = QtWidgets.QDateEdit(self.frame_7)
        self.dateEdit_4.setDate(QtCore.QDate(datetime.date.today().year, 1, 1))
        self.dateEdit_4.setObjectName("dateEdit_4")
        self.horizontalLayout_3.addWidget(self.dateEdit_4)
        # ComboBox_3 in Frame_7
        self.comboBox_3 = QtWidgets.QComboBox(self.frame_7)
        self.comboBox_3.setObjectName("comboBox_3")
        self.horizontalLayout_3.addWidget(self.comboBox_3)
        self.comboBox_3.addItem("January")
        self.comboBox_3.addItem("February")
        self.comboBox_3.addItem("March")
        self.comboBox_3.addItem("April")
        self.comboBox_3.addItem("May")
        self.comboBox_3.addItem("June")
        self.comboBox_3.addItem("July")
        self.comboBox_3.addItem("August")
        self.comboBox_3.addItem("September")
        self.comboBox_3.addItem("October")
        self.comboBox_3.addItem("November")
        self.comboBox_3.addItem("December")
        self.comboBox_3.setCurrentText(num_to_month_full(datetime.date.today().month))
        self.comboBox_3.adjustSize()
        # PushButton_5 in  frame_6
        self.frame_8 = QtWidgets.QFrame(self.frame_6)
        self.frame_8.setGeometry(QtCore.QRect(590, 10, 250, 40))
        self.pushButton_5 = QtWidgets.QPushButton(self.frame_8)
        self.pushButton_5.setGeometry(QtCore.QRect(100, 10, 70, 20))
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_5.setText("Execute")
        self.pushButton_5.clicked.connect(self.monthly_spending)
        # retranslateUi
        self.retranslateUi(Dialog)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Account Book", "Account Book"))
        self.label.setText(_translate("Account Book", "Folder Name:"))
        self.pushButton.setText(_translate("Account Book", "Load"))
        self.pushButton_2.setText(_translate("Account Book", "Edit Enabled"))
        self.pushButton_3.setText(_translate("Account Book", "New Category"))
        self.pushButton_4.setText(_translate("Account Book", "Save Data"))
        self.label_3.setText(_translate("Account Book", "Month"))
        self.dateEdit.setDisplayFormat(_translate("Account Book", "yyyy"))
        self.label_2.setText(_translate("Account Book", "Year"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Account Book", "Monthly Spending"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Account Book", "Monthly Spending"))
        self.label_6.setText(_translate("Account Book", "From: "))
        self.dateEdit_3.setDisplayFormat(_translate("Account Book", "yyyy"))
        self.label_7.setText(_translate("Account Book", "To:"))
        self.dateEdit_4.setDisplayFormat(_translate("Account Book", "yyyy"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab2), _translate("Account Book", "Yearly Spending"))

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())
